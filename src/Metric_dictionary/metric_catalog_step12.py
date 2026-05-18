"""
metric_catalog.py
─────────────────
Stage 2 — Metric Catalog Generator

PURPOSE:
    For every in-scope measure, extract:
        - measure_name, DAX, SQL
        - tables used (Snowflake objects from sf_refs)
        - columns used (Snowflake columns from sf_refs)
        - relationships (join_paths)
        - upstream dependencies

    Then call LLM to generate:
        - technical_definition : SQL/DAX-level explanation (for engineers)
        - business_definition  : plain-English meaning       (for analysts/business)

OUTPUT:
    output/dashboards/<dash>/stage2/metric_catalog.json   — full structured data
    output/dashboards/<dash>/stage2/metric_catalog.md     — markdown table

USAGE:
    python metric_catalog.py                          # pac-dash (default)
    python metric_catalog.py --dashboard risk-dash
    python metric_catalog.py --dashboard pac-dash --skip-registry
    python metric_catalog.py --measure "PAC PMPM"
    python metric_catalog.py --dry-run
"""

from __future__ import annotations
import json
import os
import sys
import argparse
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    from dotenv import load_dotenv
    _env = Path(__file__).resolve().parent
    for _ in range(4):
        c = _env / ".env"
        if c.exists():
            load_dotenv(c)
            break
        _env = _env.parent
    else:
        load_dotenv()
except ImportError:
    pass

try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

WORKERS = 5

# ══════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════

BASE_DIR = Path(__file__).resolve().parent.parent.parent

# Hardcoded per-dashboard configs removed — resolved dynamically from dashboard name.
# DASHBOARD_CONFIGS = {"pac-dash": {...}, "risk-dash": {...}}  # was: hardcoded
DASHBOARD_CONFIGS = {}  # kept for backward compat; all dashboards resolved dynamically below

def _dash_catalog_cfg(dashboard: str) -> dict:
    """Return path config for any dashboard — no hardcoding needed."""
    out = BASE_DIR / "output" / "dashboards" / dashboard / "metric_dictionary"
    return {
        "llm_json"  : out / "final_measures_with_llm.json",
        "output_dir": out,
    }

# ══════════════════════════════════════════════════════════════
# LLM CLIENT
# ══════════════════════════════════════════════════════════════

def get_client() -> OpenAI:
    base_url = os.getenv("TF_BASE_URL")
    api_key  = os.getenv("TF_API_KEY")
    model    = os.getenv("TF_MODEL")
    missing  = [k for k, v in [("TF_BASE_URL", base_url), ("TF_API_KEY", api_key), ("TF_MODEL", model)] if not v]
    if missing:
        print(f"\n[ERROR] Missing env vars: {missing}")
        sys.exit(1)
    return OpenAI(base_url=base_url, api_key=api_key)

# ══════════════════════════════════════════════════════════════
# PROMPTS
# ══════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are a healthcare analytics expert who writes concise metric definitions.

Given a Power BI DAX measure with its Snowflake SQL, tables, and columns:
  - Write a TECHNICAL definition: what the SQL/DAX computes, which tables/columns are used, how the logic works
  - Write a BUSINESS definition: what this metric means to a healthcare analyst or business user, no SQL/code

Rules:
  - Technical: 2-3 sentences max. Mention key columns, filters, aggregation type.
  - Business: 1-2 sentences max. Plain English. No column names, no SQL.
  - Respond ONLY in this exact JSON format (no markdown, no extra text):

{
  "technical": "...",
  "business": "..."
}"""


def _build_prompt(entry: dict) -> str:
    name        = entry["measure_name"]
    dax         = entry.get("dax") or entry.get("clean_dax") or entry.get("raw_dax") or ""
    sql         = entry.get("sql") or ""
    tables      = entry.get("tables") or []
    columns     = entry.get("columns") or []
    rels        = entry.get("relationships") or []
    deps        = entry.get("dependencies") or []
    pattern     = entry.get("dax_pattern") or "UNKNOWN"

    llm_defn = entry.get("llm_definition") or ""

    parts = [
        f"Measure: {name}",
        f"DAX pattern: {pattern}",
        f"DAX expression:\n{dax}",
    ]
    if sql:
        parts.append(f"SQL:\n{sql}")
    if tables:
        parts.append(f"Tables used: {', '.join(tables)}")
    if columns:
        parts.append(f"Columns used: {', '.join(columns)}")
    if rels:
        parts.append(f"Table relationships: {', '.join(rels)}")
    if deps:
        parts.append(f"Depends on measures: {', '.join(deps)}")
    if llm_defn and not sql:
        # No SQL available — give the LLM a prior definition to anchor the response
        parts.append(f"Prior definition (use as context):\n{llm_defn}")

    return "\n\n".join(parts)


# ══════════════════════════════════════════════════════════════
# EXTRACTION
# ══════════════════════════════════════════════════════════════

def _collect_sf_refs(m: dict, lookup: dict, visited: set) -> list:
    """Recursively collect sf_refs from a measure and all its dependencies."""
    name = m.get("measure_name", "")
    if name in visited:
        return []
    visited.add(name)

    refs = list(m.get("sf_refs") or [])
    for dep in (m.get("depends_on") or []):
        dep_name = dep.get("measure_name", "")
        dep_full = lookup.get(dep_name)
        if dep_full:
            refs.extend(_collect_sf_refs(dep_full, lookup, visited))
    return refs


def _extract_dax_refs(dax: str) -> tuple:
    """
    Fallback: extract table[column] pairs directly from raw DAX.
    Used when sf_refs is empty (RUNTIME_ROUTER / DEFINER scope — no SQL generated).
    Returns (tables, columns) as sorted lists of BI names.
    """
    import re
    # Match table[column] — skip pure measure references like [Measure Name] (no table prefix)
    pairs = re.findall(r"'?([\w][\w\s]*)'?\[([\w][\w\s]*)\]", dax)
    tables  = sorted({t.strip() for t, _ in pairs})
    columns = sorted({c.strip() for _, c in pairs})
    return tables, columns


def extract_entry(m: dict, lookup: Optional[dict] = None) -> dict:
    """Extract structured fields from a measure record."""
    all_refs = _collect_sf_refs(m, lookup or {}, set())

    tables  = sorted({r["sf_object"] for r in all_refs if r.get("sf_object") and r.get("ref_type") == "source"})
    columns = sorted({
        r["sf_column"]
        for r in all_refs
        if r.get("sf_object") and r.get("sf_column") and r.get("ref_type") == "source"
    })

    # No Snowflake refs (measure was skipped by compiler) — fall back to DAX table[column] refs
    if not tables and not columns:
        dax = (m.get("clean_dax") or m.get("raw_dax") or "").strip()
        tables, columns = _extract_dax_refs(dax)

    rels    = m.get("join_paths") or []
    deps    = [d["measure_name"] for d in (m.get("depends_on") or [])]
    sql     = m.get("sql_query") or ""

    return {
        "measure_name"  : m["measure_name"],
        "dax_pattern"   : m.get("dax_pattern"),
        "scope"         : m.get("scope"),
        "dax"           : (m.get("clean_dax") or m.get("raw_dax") or "").strip(),
        "sql"           : sql.strip(),
        "tables"        : tables,
        "columns"       : columns,
        "relationships" : rels,
        "dependencies"  : deps,
        "llm_definition": m.get("llm_definition") or "",
        "technical_definition": None,
        "business_definition" : None,
    }


# ══════════════════════════════════════════════════════════════
# LLM CALL
# ══════════════════════════════════════════════════════════════

_print_lock = threading.Lock()

def _define_one(
    idx      : int,
    total    : int,
    entry    : dict,
    client   : OpenAI,
    registry : dict,
    dry_run  : bool,
    skip_reg : bool,
) -> dict:
    name = entry["measure_name"]

    # Registry cache check
    cached = registry.get(name)
    if cached and not skip_reg:
        with _print_lock:
            print(f"  [{idx:3d}/{total}] {name:<50s} 📋 CACHED")
        entry["technical_definition"] = cached.get("technical_definition")
        entry["business_definition"]  = cached.get("business_definition")
        return entry

    if dry_run:
        with _print_lock:
            print(f"  [{idx:3d}/{total}] {name:<50s} 🔍 DRY_RUN")
        return entry

    user_prompt = _build_prompt(entry)
    model       = os.getenv("TF_MODEL")

    import re, time as _time
    last_exc = None
    for _attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model,
                max_completion_tokens=600,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_prompt},
                ],
            )
            content = resp.choices[0].message.content
            finish  = resp.choices[0].finish_reason

            if not content or content.strip().startswith("ERROR:"):
                raise ValueError(f"Empty response (finish_reason={finish})")

            clean = content.strip()
            fence = re.search(r"```(?:json)?\s*\n([\s\S]*?)(?:\n```|$)", clean)
            if fence:
                clean = fence.group(1).strip()

            if not clean:
                raise ValueError(f"Empty JSON inside fence (finish_reason={finish})")

            result = json.loads(clean)
            entry["technical_definition"] = result.get("technical", "")
            entry["business_definition"]  = result.get("business",  "")

            with _print_lock:
                print(f"  [{idx:3d}/{total}] {name:<50s} ✅ DEFINED")
            last_exc = None
            break

        except Exception as exc:
            last_exc = exc
            if _attempt < 2:
                _time.sleep(2 ** _attempt)

    if last_exc is not None:
        entry["technical_definition"] = f"ERROR: {last_exc}"
        entry["business_definition"]  = f"ERROR: {last_exc}"
        with _print_lock:
            print(f"  [{idx:3d}/{total}] {name:<50s} ⚠️  ERROR — {str(last_exc)[:60]}")

    return entry


# ══════════════════════════════════════════════════════════════
# MARKDOWN TABLE
# ══════════════════════════════════════════════════════════════

def _to_markdown(entries: list[dict]) -> str:
    headers = ["Measure Name", "Dax", "Tables", "Columns", "SQL", "Technical Definition", "Business Definition"]
    rows = []
    for e in entries:
        rows.append([
            e["measure_name"],
            e.get("dax") or "",
            ", ".join(e.get("tables") or []),
            ", ".join(e.get("columns") or []),
            ", ".join(e.get("sql") or []),
            (e.get("technical_definition") or "").replace("\n", " "),
            (e.get("business_definition")  or "").replace("\n", " "),
        ])

    # Column widths
    widths = [max(len(h), max((len(r[i]) for r in rows), default=0)) for i, h in enumerate(headers)]
    widths = [min(w, 80) for w in widths]  # cap at 80

    def _row(cells):
        return "| " + " | ".join(str(c)[:80].ljust(w) for c, w in zip(cells, widths)) + " |"

    sep = "|-" + "-|-".join("-" * w for w in widths) + "-|"
    lines = [_row(headers), sep] + [_row(r) for r in rows]
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

EXCEL_COLUMNS = [
    ("measure_name",         "Measure Name"),
    ("dax",                  "DAX"),
    ("sql",                  "SQL"),
    ("tables",               "Tables"),
    ("columns",              "Columns"),
    ("technical_definition", "Technical Definition"),
    ("business_definition",  "Business Definition"),
]

def _to_excel(entries: list[dict], out_path: Path) -> None:
    if not PANDAS_AVAILABLE:
        print("  ⚠ pandas not installed — skipping Excel. Run: pip install pandas openpyxl")
        return

    rows = []
    for e in entries:
        def _fmt(key, val):
            if not isinstance(val, list):
                return val or ""
            if key == "columns":
                return ", ".join(c.split(".")[-1] for c in val)
            return ", ".join(val)

        rows.append({label: _fmt(key, e.get(key)) for key, label in EXCEL_COLUMNS})

    df = pd.DataFrame(rows, columns=[label for _, label in EXCEL_COLUMNS])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Metric Catalog")
        ws = writer.sheets["Metric Catalog"]

        # Column widths
        col_widths = {"Measure Name": 30, "DAX": 40, "SQL": 50,
                      "Tables": 30, "Columns": 40,
                      "Technical Definition": 60, "Business Definition": 60}
        for col_idx, (_, label) in enumerate(EXCEL_COLUMNS, start=1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = col_widths.get(label, 20)

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="2E2E2E")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Row styling
        alt_fill = PatternFill("solid", fgColor="F5F5F5")
        for row in ws.iter_rows(min_row=2):
            fill = alt_fill if row[0].row % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(name="Calibri", size=10)

    print(f"    {out_path}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def run_catalog(
    dashboard    : str  = "pac-dash",
    measure_filter: Optional[str] = None,
    dry_run      : bool = False,
    skip_registry: bool = False,
    scope_filter : str  = "all",
    excel        : bool = False,
) -> None:
    cfg        = _dash_catalog_cfg(dashboard)
    llm_json   = cfg["llm_json"]
    output_dir = cfg["output_dir"]
    reg_path   = output_dir / "metric_catalog_registry.json"
    out_json   = output_dir / "metric_catalog.json"
    out_md     = output_dir / "metric_catalog.md"

    if not llm_json.exists():
        print(f"\n❌ {llm_json} not found. Run llm_fallback.py first.")
        sys.exit(1)

    try:
        measures = json.loads(llm_json.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"\n❌ {llm_json} is malformed ({e}). Re-run llm_fallback.py.")
        sys.exit(1)
    try:
        registry = json.loads(reg_path.read_text(encoding="utf-8")) if reg_path.exists() else {}
    except json.JSONDecodeError as e:
        print(f"[metric_catalog] WARNING: registry.json is malformed ({e}); defaulting to {{}}.")
        registry = {}

    print("=" * 60)
    print(f"  Metric Catalog — {dashboard}")
    print("=" * 60)
    print(f"\n  Loaded {len(measures)} measures")

    # Build name -> measure lookup for dependency walking
    measures_lookup = {m["measure_name"]: m for m in measures}

    # Extract structured entries
    entries = []
    for m in measures:
        if measure_filter and m["measure_name"] != measure_filter:
            continue
        if scope_filter != "all" and m.get("scope") != scope_filter:
            continue
        entries.append(extract_entry(m, measures_lookup))

    print(f"  Processing {len(entries)} measures (scope={scope_filter})")
    if dry_run:
        print("  DRY RUN — no API calls\n")

    client = get_client() if (OPENAI_AVAILABLE and not dry_run) else None

    # Run LLM concurrently
    results = [None] * len(entries)
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = {
            pool.submit(
                _define_one,
                idx + 1, len(entries), entry, client, registry, dry_run, skip_registry
            ): idx
            for idx, entry in enumerate(entries)
        }
        for fut in as_completed(futures):
            idx = futures[fut]
            results[idx] = fut.result()

    # Update registry
    now = datetime.now(timezone.utc).isoformat()
    for r in results:
        if r and r.get("technical_definition") and not str(r["technical_definition"]).startswith("ERROR"):
            registry[r["measure_name"]] = {
                "technical_definition": r["technical_definition"],
                "business_definition" : r["business_definition"],
                "updated_at"          : now,
            }

    # Write outputs
    output_dir.mkdir(parents=True, exist_ok=True)
    reg_path.write_text(json.dumps(registry, indent=2, ensure_ascii=False), encoding="utf-8")
    out_json.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    out_md.write_text(_to_markdown(results), encoding="utf-8")

    if excel:
        out_xlsx = output_dir / "metric_catalog.xlsx"
        _to_excel(results, out_xlsx)

    defined  = sum(1 for r in results if r and r.get("technical_definition") and not str(r.get("technical_definition","")).startswith("ERROR"))
    errors   = sum(1 for r in results if r and str(r.get("technical_definition","")).startswith("ERROR"))
    cached_n = sum(1 for r in results if r and registry.get(r["measure_name"]) and not skip_registry)

    print(f"\n{'─'*60}")
    print(f"  COMPLETE")
    print(f"{'─'*60}")
    print(f"  Defined  : {defined}")
    print(f"  Cached   : {cached_n}")
    print(f"  Errors   : {errors}")
    print(f"\n  Output:")
    print(f"    {out_json}")
    print(f"    {out_md}")
    if excel:
        print(f"    {output_dir / 'metric_catalog.xlsx'}")
    print(f"    {reg_path}")


# ══════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Metric Catalog — technical + business definitions")
    parser.add_argument("--dashboard",    type=str, default="risk-dash",  help="pac-dash | risk-dash (default: pac-dash)")
    parser.add_argument("--measure",      type=str, default=None,        help="Process only this measure name")
    parser.add_argument("--scope",        type=str, default="all",       help="IN_SCOPE | OUT_OF_SCOPE | all (default: all)")
    parser.add_argument("--dry-run",      action="store_true",           help="Print without calling LLM")
    parser.add_argument("--skip-registry",action="store_true",           help="Ignore cache, re-generate all")
    parser.add_argument(
    "--excel",
    action="store_true",
    default=True,
    help="Also write metric_catalog.xlsx"
)
    args = parser.parse_args()

    run_catalog(
        dashboard     = args.dashboard,
        measure_filter= args.measure,
        dry_run       = args.dry_run,
        skip_registry = args.skip_registry,
        scope_filter  = args.scope,
        excel         = args.excel,
    )
